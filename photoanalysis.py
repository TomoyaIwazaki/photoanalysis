# -*- coding: utf-8 -*-
import os
import cv2
import glob
import numpy as np
import openpyxl
from matplotlib import pyplot as plt


def main():
    #画像を保存したファイル名を入力
    data_dir_path = u"./photodata1/"
    file_list = os.listdir(r"./photodata1/")
    
    #エクセルのファイル名とシート名を入力
    wd = openpyxl.load_workbook('解析データ.xlsx')
    ws = wd['Sheet1']
    
    ws.append(['red', 'red-orange', 'yellow-orange',
               'yellow', 'yellow-grren', 'green',
               'blue-green', 'green-blue', 'blue',
               'blue-violet', 'violet', 'red-violet','white','black'])

    for file_name in file_list:
        root, ext = os.path.splitext(file_name)
        if ext == u'.png' or u'.jpeg' or u'.jpg':
            abs_name = data_dir_path + '/' + file_name
            img = cv2.imread(abs_name, cv2.IMREAD_COLOR)
            imghsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
            
        # 以下各画像に対する処理を記載する

        color_wheel = ['red', 'red-orange', 'yellow-orange',
                      'yellow', 'yellow-grren', 'green',
                      'blue-green', 'green-blue', 'blue',
                      'blue-violet', 'violet', 'red-violet', 'red',
                      'white', 'black']

        color_count = {
                       'red': 0,
                       'red-orange': 0,
                       'yellow-orange': 0,
                       'yellow': 0,
                       'yellow-grren': 0,
                       'green': 0,
                       'blue-green': 0,
                       'green-blue': 0,
                       'blue': 0,
                       'blue-violet': 0,
                       'violet': 0,
                       'red-violet': 0,
                       'white':0,
                       'black':0
                     }
        #色判定   
        def color_judge(value):
            index_value = int(round(value/(180/(len(color_wheel)-3)), 0))
            return color_wheel[index_value]

# 選択した場所の色を出力する

        for v in imghsv:
            for h in v:
                if((h[1]) > 30):
                    if((h[2]) > 30):
                        color_index = color_judge(h[0])
                        color_count[color_index] += 1
                    else:
                        color_count['black'] += 1
                else:
                    color_count['white'] += 1

        a = ws.max_row
        x = 65
        for i in color_count.values():
            ws[chr(x) + str(a+1)] = i 
            x += 1
    
    wd.save('rogoデータ.xlsx')
    wd.close()
    
if __name__ == '__main__':
    main()
